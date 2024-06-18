import json
from openpyxl import load_workbook

def convert_xlsx_to_txt(input_xlsx_path, output_txt_path):
    # 加载 Excel 文件
    workbook = load_workbook(filename=input_xlsx_path, read_only=True)
    sheet = workbook.active

    # 获取初始行数
    initial_line_count = sheet.max_row
    print(f"Initial line count: {initial_line_count}")

    # 打开输出文件进行写入
    with open(output_txt_path, 'w', encoding='utf-8') as f:
        for row in sheet.iter_rows(values_only=True):
            # 将每行的第一个单元格值写入到文本文件
            item = row[0]
            if item is not None:
                # 将每个单元格内容作为 JSON 字符串写入文件，以确保换行符不会影响行数
                json_item = json.dumps(str(item))
                f.write(f"{json_item}\n")
            else:
                f.write("\n")

    # 验证输出文件的行数
    with open(output_txt_path, 'r', encoding='utf-8') as f:
        output_line_count = sum(1 for line in f)

    print(f"Output file line count: {output_line_count}")

    assert initial_line_count == output_line_count, (
        f"Line count mismatch: Excel file has {initial_line_count} lines, "
        f"but output file has {output_line_count} lines."
    )

    print(f"Conversion complete. {initial_line_count} lines written to {output_txt_path}")

if __name__ == "__main__":
    input_xlsx_path = 'documents.xlsx'  # 替换为你的 Excel 文件路径
    output_txt_path = 'documents.txt'  # 替换为输出文本文件路径
    convert_xlsx_to_txt(input_xlsx_path, output_txt_path)
